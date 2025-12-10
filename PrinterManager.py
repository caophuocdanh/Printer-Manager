# --- START OF FILE PrinterManager.py ---

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import winreg
import subprocess
import ctypes
import os
import sys
import threading
import datetime
import re
import csv
import shutil

# --- HÀM HỖ TRỢ HỆ THỐNG ---
def resource_path(relative_path):
    try: base_path = sys._MEIPASS
    except: base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def is_admin():
    try: return ctypes.windll.shell32.IsUserAnAdmin()
    except: return False

def run_as_admin():
    try:
        if getattr(sys, 'frozen', False):
            ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, None, None, 1)
        else:
            ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, f'"{os.path.abspath(sys.argv[0])}"', None, 1)
        return True
    except: return False

# --- CẤU HÌNH & HẰNG SỐ ---
APP_VERSION = "2.2.1"
APP_BUILD = "Final_Polished"
APP_AUTHOR = "@danhcp"
APP_TITLE = "PRINTER MANAGER PRO"
ICON_NAME = "printer.ico"

# --- BẢNG MÀU HIỆN ĐẠI (FLAT DESIGN) ---
C_SIDEBAR_BG    = "#2C3E50"  # Xanh đen đậm
C_SIDEBAR_FG    = "#ECF0F1"  # Trắng đục
C_MAIN_BG       = "#ECF0F1"  # Xám rất nhạt
C_CARD_BG       = "#FFFFFF"  # Trắng tinh
C_ACCENT        = "#2980B9"  # Xanh dương đậm
C_ACCENT_HOVER  = "#3498DB"  # Xanh dương sáng
C_TEXT_DARK     = "#2C3E50"  # Màu chữ chính
C_TEXT_LIGHT    = "#7F8C8D"  # Màu chữ phụ
C_DANGER        = "#E74C3C"  # Màu đỏ

REG_PRINTERS = r"SYSTEM\CurrentControlSet\Control\Print\Printers"
REG_DRIVERS_V3 = r"SYSTEM\CurrentControlSet\Control\Print\Environments\Windows x64\Drivers\Version-3"
REG_DRIVERS_V4 = r"SYSTEM\CurrentControlSet\Control\Print\Environments\Windows x64\Drivers\Version-4"
SPOOL_DIR = r"C:\Windows\System32\spool\PRINTERS"
BACKUP_DIR = "Backup"
LOG_FILE = "activity.log"
PRINTER_ATTRIBUTE_SHARED = 0x00000008

class CleanPrinterApp:
    def __init__(self, root):
        self.root = root
        self.setup_window()
        self.setup_styles()
        self.setup_layout()
        self.create_context_menu()
        
        # Init logic
        self.log(f"System Ready - {APP_TITLE} v{APP_VERSION}")
        self.check_spooler_status_on_startup()
        self.scan_printers() # Auto scan khi mở

    def setup_window(self):
        try: ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(f'danhcp.printermanager.{APP_BUILD}')
        except: pass
        try: self.root.iconbitmap(resource_path(ICON_NAME))
        except: pass
        
        self.root.title(f"{APP_TITLE}")
        self.root.geometry("1150x720")
        self.root.minsize(1000, 600)
        self.root.configure(bg=C_MAIN_BG)

    def setup_styles(self):
        self.style = ttk.Style()
        self.style.theme_use('clam') 

        # Style cho Treeview (Bảng dữ liệu)
        self.style.configure("Treeview",
                             background=C_CARD_BG,
                             foreground=C_TEXT_DARK,
                             fieldbackground=C_CARD_BG,
                             font=("Segoe UI", 10),
                             rowheight=35, 
                             borderwidth=0)
        
        self.style.map('Treeview', background=[('selected', C_ACCENT)])

        # Header của bảng
        self.style.configure("Treeview.Heading",
                             background="#DADFE1",
                             foreground="#2C3E50",
                             font=("Segoe UI", 9, "bold"),
                             relief="flat")
        
        # Scrollbar (Đậm hơn để dễ nhìn)
        self.style.configure("Vertical.TScrollbar", gripcount=0, background="#95A5A6", troughcolor="#ECF0F1", borderwidth=0, arrowsize=14)
        self.style.configure("Horizontal.TScrollbar", gripcount=0, background="#95A5A6", troughcolor="#ECF0F1", borderwidth=0, arrowsize=14)

    def setup_layout(self):
        self.root.columnconfigure(0, weight=0) # Sidebar cố định
        self.root.columnconfigure(1, weight=1) # Content giãn
        self.root.rowconfigure(0, weight=1)

        # === 1. SIDEBAR (MENU TRÁI) ===
        self.sidebar = tk.Frame(self.root, bg=C_SIDEBAR_BG, width=250)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.pack_propagate(False)

        # Logo
        lbl_title = tk.Label(self.sidebar, text="PRINTER\nMANAGER", bg=C_SIDEBAR_BG, fg="white", font=("Segoe UI", 16, "bold"), justify="left", pady=25, padx=20)
        lbl_title.pack(anchor="w")

        # Menu Wrapper
        menu_frame = tk.Frame(self.sidebar, bg=C_SIDEBAR_BG, padx=10)
        menu_frame.pack(fill="x", pady=5)

        # Các nút chức năng Sidebar
        self.btn_sidebar(menu_frame, "🔄  Quét / Làm Mới", self.scan_printers, primary=True)
        self.create_spacer(menu_frame)
        self.btn_sidebar(menu_frame, "➕  Thêm Máy In", self.action_add_printer)
        self.btn_sidebar(menu_frame, "🧹  Dọn Driver Rác", self.action_delete_unused_drivers)
        self.create_spacer(menu_frame)
        self.btn_sidebar(menu_frame, "♻️  Restart Spooler", lambda: self.run_thread(self.restart_spooler))
        self.btn_sidebar(menu_frame, "🔥  Xóa Lệnh In", self.clear_spool_files) # Đã đổi tên nút
        self.create_spacer(menu_frame)
        self.btn_sidebar(menu_frame, "📂  Print Management", lambda: self.run_cmd("printmanagement.msc"))
        self.btn_sidebar(menu_frame, "⚙️  Control Panel", lambda: self.run_cmd("control printers"))
        self.btn_sidebar(menu_frame, "📊  Xuất Excel (.xlsx)", self.export_report)

        # Footer Sidebar
        lbl_ver = tk.Label(self.sidebar, text=f"v{APP_VERSION}\nDev: {APP_AUTHOR}", bg=C_SIDEBAR_BG, fg="#7F8C8D", font=("Segoe UI", 8), pady=20)
        lbl_ver.pack(side="bottom")


        # === 2. MAIN CONTENT ===
        self.content = tk.Frame(self.root, bg=C_MAIN_BG)
        self.content.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        self.content.rowconfigure(1, weight=1) # Row chứa bảng giãn nở
        self.content.columnconfigure(0, weight=1)

        # Header
        top_bar = tk.Frame(self.content, bg=C_MAIN_BG)
        top_bar.grid(row=0, column=0, sticky="ew", pady=(0, 15))
        tk.Label(top_bar, text="Danh Sách Máy In Hệ Thống", bg=C_MAIN_BG, fg=C_TEXT_DARK, font=("Segoe UI", 18)).pack(side="left")
        
        # Status Label
        self.lbl_spooler = tk.Label(top_bar, text="Spooler: Checking...", bg="#95A5A6", fg="white", font=("Segoe UI", 9, "bold"), padx=12, pady=5)
        self.lbl_spooler.pack(side="right")


        # === 3. CARD VIEW (BẢNG + SCROLLBAR) ===
        # Sử dụng grid layout bên trong card_frame
        card_frame = tk.Frame(self.content, bg=C_CARD_BG)
        card_frame.grid(row=1, column=0, sticky="nsew")
        card_frame.rowconfigure(0, weight=1) 
        card_frame.columnconfigure(0, weight=1)

        cols = ("no", "status", "name", "port", "driver", "share") 
        self.tree = ttk.Treeview(card_frame, columns=cols, show="headings", selectmode="browse")
        
        self.tree.heading("no", text="#"); self.tree.column("no", width=40, anchor="center")
        self.tree.heading("status", text="TRẠNG THÁI"); self.tree.column("status", width=120)
        self.tree.heading("name", text="TÊN MÁY IN"); self.tree.column("name", width=280)
        self.tree.heading("port", text="CỔNG (PORT)"); self.tree.column("port", width=150)
        self.tree.heading("driver", text="DRIVER"); self.tree.column("driver", width=200)
        self.tree.heading("share", text="CHIA SẺ"); self.tree.column("share", width=70, anchor="center")

        # Scrollbars
        vsb = ttk.Scrollbar(card_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(card_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # === 4. LOG PANEL ===
        log_frame = tk.LabelFrame(self.content, text="Nhật ký hệ thống", bg=C_MAIN_BG, fg=C_TEXT_LIGHT, font=("Segoe UI", 9), relief="flat", pady=5)
        log_frame.grid(row=2, column=0, sticky="ew", pady=(15, 0))
        
        self.txt_log = scrolledtext.ScrolledText(log_frame, height=5, state='disabled', font=("Consolas", 9), bg="white", relief="flat", padx=5, pady=5)
        self.txt_log.pack(fill="x")

        # Tags màu
        self.tree.tag_configure('default_printer', foreground=C_ACCENT, font=('Segoe UI', 10, 'bold'))
        self.tree.tag_configure('offline', foreground="#95A5A6") 
        self.tree.tag_configure('error', foreground="#E74C3C") 

        self.tree.bind("<Button-3>", self.show_context_menu)


    # --- UI HELPER ---
    def btn_sidebar(self, parent, text, cmd, primary=False):
        bg = C_ACCENT if primary else C_SIDEBAR_BG
        fg = "white" if primary else "#BDC3C7"
        font = ("Segoe UI", 10, "bold") if primary else ("Segoe UI", 10)
        
        btn = tk.Button(parent, text=text, bg=bg, fg=fg, font=font, command=cmd, 
                        relief="flat", bd=0, padx=15, pady=8, anchor="w", cursor="hand2")
        btn.pack(fill="x", pady=2)
        
        if not primary:
            def on_e(e): btn['bg'] = "#34495E"; btn['fg'] = "white"
            def on_l(e): btn['bg'] = C_SIDEBAR_BG; btn['fg'] = "#BDC3C7"
            btn.bind("<Enter>", on_e); btn.bind("<Leave>", on_l)
        else:
            def on_e(e): btn['bg'] = C_ACCENT_HOVER
            def on_l(e): btn['bg'] = C_ACCENT
            btn.bind("<Enter>", on_e); btn.bind("<Leave>", on_l)

    def create_spacer(self, parent):
        tk.Frame(parent, height=1, bg="#34495E").pack(fill="x", pady=10)


    # --- CORE LOGIC ---
    def log(self, msg):
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        try:
            self.txt_log.config(state='normal')
            self.txt_log.insert("end", f"[{timestamp}] {msg}\n")
            self.txt_log.see("end")
            self.txt_log.config(state='disabled')
        except: pass
        try:
            with open(LOG_FILE, "a", encoding="utf-8") as f:
                f.write(f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} {msg}\n")
        except: pass

    def run_cmd(self, cmd):
        try: subprocess.Popen(cmd, shell=True)
        except Exception as e: self.log(f"Err cmd: {e}")

    def run_thread(self, func, args=()):
        threading.Thread(target=func, args=args, daemon=True).start()

    def check_spooler_status_on_startup(self):
        def _check():
            try:
                cmd = 'powershell "Get-Service spooler | Select-Object -ExpandProperty Status"'
                si = subprocess.STARTUPINFO(); si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                out = subprocess.check_output(cmd, shell=True, startupinfo=si).decode().strip()
                color = "#27AE60" if out == "Running" else "#E74C3C"
                self.lbl_spooler.config(text=f"Spooler: {out}", bg=color)
            except Exception as e: self.log(f"Spooler check fail: {e}")
        self.run_thread(_check)

    # --- SCANNING ---
    def get_default_printer_name(self):
        try:
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows NT\CurrentVersion\Windows") as key:
                val, _ = winreg.QueryValueEx(key, "Device")
                return val.split(',')[0]
        except: return None

    def get_printer_statuses_map(self):
        status_map = {}
        try:
            cmd = 'powershell "Get-Printer | Select-Object Name, PrinterStatus | ConvertTo-Csv -NoTypeInformation"'
            si = subprocess.STARTUPINFO(); si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            process = subprocess.Popen(cmd, stdout=subprocess.PIPE, text=True, startupinfo=si)
            out, _ = process.communicate()
            lines = out.strip().splitlines()
            if len(lines) > 1:
                reader = csv.reader(lines); next(reader)
                for row in reader:
                    if len(row) >= 2: status_map[row[0]] = row[1]
        except: pass
        return status_map

    def translate_status(self, status_str):
        s = status_str.lower()
        if s == "normal" or s == "idle": return "🟢 Sẵn sàng"
        if s == "printing": return "🖨️ Đang in"
        if s == "paused": return "⏸️ Tạm dừng"
        if s in ["error", "drivererror"]: return "🔴 Lỗi"
        if s == "offline": return "⚫ Offline"
        return f"⚪ {status_str}"

    def scan_printers(self):
        self.run_thread(self._scan_printers_worker)

    def _scan_printers_worker(self):
        self.root.after(0, lambda: [self.tree.delete(item) for item in self.tree.get_children()])
        self.log("⏳ Đang tải dữ liệu...")
        
        status_map = self.get_printer_statuses_map()
        default_printer = self.get_default_printer_name()
        
        try:
            hKey = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, REG_PRINTERS)
            idx = 0; count = 0
            items = []

            while True:
                try:
                    p_name = winreg.EnumKey(hKey, idx)
                    tags = []
                    display_name = p_name
                    if p_name == default_printer:
                        display_name = f"⭐ {p_name}"
                        tags.append('default_printer')

                    raw_status = status_map.get(p_name, "Unknown")
                    display_status = self.translate_status(raw_status)
                    if "Offline" in display_status: tags.append('offline')
                    if "Lỗi" in display_status: tags.append('error')

                    d_name = "N/A"; port_name = "N/A"; share_status = "-"
                    try:
                        sub = winreg.OpenKey(hKey, p_name)
                        try: d_name, _ = winreg.QueryValueEx(sub, "Printer Driver")
                        except: pass
                        try: port_name, _ = winreg.QueryValueEx(sub, "Port")
                        except: pass
                        try: 
                            attr, _ = winreg.QueryValueEx(sub, "Attributes")
                            if attr & PRINTER_ATTRIBUTE_SHARED: share_status = "✅"
                        except: pass
                        winreg.CloseKey(sub)
                    except: pass
                    
                    count += 1
                    items.append({
                        'values': (count, display_status, display_name, port_name, d_name, share_status),
                        'tags': tuple(tags)
                    })
                    idx += 1
                except OSError: break
            winreg.CloseKey(hKey)
            
            self.root.after(0, lambda: [self.tree.insert("", "end", values=i['values'], tags=i['tags']) for i in items])
            self.root.after(0, lambda: self.log(f"✅ Đã tải xong {count} máy in."))

        except Exception as e: self.log(f"Lỗi: {e}")

    # --- DRIVER CLEANUP ---
    def action_delete_unused_drivers(self):
        self.log("🧹 Đang quét driver...")
        self.run_thread(self._find_unused_drivers_thread)

    def _find_unused_drivers_thread(self):
        try:
            in_use = set()
            hKey_printers = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, REG_PRINTERS)
            idx = 0
            while True:
                try:
                    sub = winreg.OpenKey(hKey_printers, winreg.EnumKey(hKey_printers, idx))
                    try:
                        dn, _ = winreg.QueryValueEx(sub, "Printer Driver")
                        if dn != "N/A": in_use.add(dn)
                    except: pass
                    winreg.CloseKey(sub)
                    idx += 1
                except: break
            winreg.CloseKey(hKey_printers)

            all_drivers = set()
            for path in [REG_DRIVERS_V3, REG_DRIVERS_V4]:
                try:
                    hKey = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, path)
                    idx = 0
                    while True:
                        try: all_drivers.add(winreg.EnumKey(hKey, idx)); idx+=1
                        except: break
                    winreg.CloseKey(hKey)
                except: pass

            unused = sorted(list(all_drivers - in_use))
            self.root.after(0, self._show_unused_driver_dialog, unused)
        except Exception as e: self.log(f"Lỗi quét: {e}")

    def _show_unused_driver_dialog(self, unused_drivers):
        if not unused_drivers:
            messagebox.showinfo("Tuyệt vời", "Không tìm thấy driver thừa nào!")
            return
        
        top = tk.Toplevel(self.root)
        top.title("Dọn dẹp Driver")
        top.geometry("600x450")
        top.configure(bg=C_MAIN_BG)
        
        tk.Label(top, text=f"Tìm thấy {len(unused_drivers)} driver không sử dụng:", bg=C_MAIN_BG, font=("Segoe UI", 11)).pack(pady=10)
        
        f_list = tk.Frame(top, bg="white", bd=1, relief="solid")
        f_list.pack(fill="both", expand=True, padx=20, pady=5)
        
        lb = tk.Listbox(f_list, selectmode="extended", font=("Segoe UI", 10), bd=0, highlightthickness=0)
        sb = ttk.Scrollbar(f_list, orient="vertical", command=lb.yview)
        lb.config(yscrollcommand=sb.set)
        
        lb.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        sb.pack(side="right", fill="y")
        
        for d in unused_drivers: lb.insert("end", d)
        
        def do_del():
            sels = [lb.get(i) for i in lb.curselection()]
            if not sels: return
            if messagebox.askyesno("Cảnh báo", f"Xóa {len(sels)} driver?\nSpooler sẽ khởi động lại."):
                self.run_thread(self._process_delete_drivers, args=(sels,))
                top.destroy()
        
        btn_del = tk.Button(top, text="Xóa Mục Đã Chọn", bg=C_DANGER, fg="white", font=("Segoe UI", 10, "bold"), 
                            command=do_del, relief="flat", padx=20, pady=8, cursor="hand2")
        btn_del.pack(pady=15)

    def _process_delete_drivers(self, drivers):
        self.stop_spooler()
        for d in drivers: self.delete_driver_reg(d)
        self.start_spooler()
        self.log(f"Đã xóa {len(drivers)} driver rác.")
        self.root.after(1000, self.scan_printers)

    # --- CONTEXT MENU & ACTIONS ---
    def create_context_menu(self):
        self.context_menu = tk.Menu(self.root, tearoff=0, bg="white", fg=C_TEXT_DARK, font=("Segoe UI", 10))
        self.context_menu.add_command(label="⭐ Đặt làm Mặc định", command=self.set_default_printer)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="📄 Xem lệnh in (Queue)", command=self.view_print_queue)
        self.context_menu.add_command(label="🖨️ In trang test", command=self.action_print_test)
        self.context_menu.add_command(label="⚙️ Preferences...", command=self.open_printing_preferences)
        self.context_menu.add_command(label="🔧 Properties...", command=self.open_printer_properties)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="🌐 Ping IP", command=lambda: self.run_thread(self.action_ping))
        self.context_menu.add_command(label="🔄 Chia sẻ LAN (On/Off)", command=self.toggle_sharing)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="❌ Xóa Máy In...", command=self.action_delete)

    def show_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)

    def action_add_printer(self): self.run_cmd('start ms-settings:printers')
    
    def open_printing_preferences(self):
        sel = self.tree.selection()
        if not sel: return
        n = self.tree.item(sel[0])['values'][2].lstrip("⭐").strip()
        self.run_cmd(f'rundll32 printui.dll,PrintUIEntry /e /n "{n}"')

    def open_printer_properties(self):
        sel = self.tree.selection()
        if not sel: return
        n = self.tree.item(sel[0])['values'][2].lstrip("⭐").strip()
        self.run_cmd(f'rundll32 printui.dll,PrintUIEntry /p /n "{n}"')

    def set_default_printer(self):
        sel = self.tree.selection()
        if not sel: return
        n = self.tree.item(sel[0])['values'][2].lstrip("⭐").strip()
        subprocess.run(f'rundll32 printui.dll,PrintUIEntry /y /n "{n}"', shell=True)
        self.scan_printers()

    def view_print_queue(self):
        sel = self.tree.selection()
        if not sel: return
        n = self.tree.item(sel[0])['values'][2].lstrip("⭐").strip()
        top = tk.Toplevel(self.root)
        top.title(f"Queue: {n}")
        top.geometry("800x400")
        
        tree = ttk.Treeview(top, columns=("id", "doc", "user", "pages", "size"), show="headings")
        tree.heading("doc", text="Tài liệu"); tree.column("doc", width=250)
        tree.heading("user", text="User"); tree.column("user", width=100)
        tree.heading("pages", text="Trang"); tree.column("pages", width=50)
        tree.pack(fill="both", expand=True)
        
        def fetch():
            cmd = f'powershell "Get-PrintJob -PrinterName \'{n}\' | Select-Object Id,DocumentName,UserName,TotalPages,JobSize | ConvertTo-Csv -NoTypeInformation"'
            try:
                si = subprocess.STARTUPINFO(); si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                out = subprocess.check_output(cmd, shell=True, startupinfo=si).decode()
                lines = out.strip().splitlines()
                if len(lines) > 1:
                    reader = csv.reader(lines); next(reader)
                    for r in reader: tree.insert("", "end", values=r)
                else: tree.insert("", "end", values=("Trống", "", "", ""))
            except: pass
        self.run_thread(fetch)

    def action_print_test(self):
        sel = self.tree.selection()
        if sel:
            n = self.tree.item(sel[0])['values'][2].lstrip("⭐").strip()
            self.run_cmd(f'rundll32 printui.dll,PrintUIEntry /k /n "{n}"')

    def action_ping(self):
        sel = self.tree.selection()
        if sel:
            p = str(self.tree.item(sel[0])['values'][3])
            ip = re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', p)
            if ip:
                ip = ip.group(1)
                self.log(f"Ping {ip}...")
                si = subprocess.STARTUPINFO(); si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                proc = subprocess.Popen(f"ping -n 2 {ip}", stdout=subprocess.PIPE, startupinfo=si)
                out, _ = proc.communicate()
                if b"TTL=" in out: messagebox.showinfo("Ping", f"{ip} Online ✅")
                else: messagebox.showerror("Ping", f"{ip} Unreachable ❌")
            else: messagebox.showwarning("Lỗi", "Không tìm thấy IP trong cổng này.")

    def toggle_sharing(self):
        sel = self.tree.selection()
        if not sel: return
        n = self.tree.item(sel[0])['values'][2].lstrip("⭐").strip()
        if messagebox.askyesno("Share", f"Đổi trạng thái Share cho {n}?"):
            try:
                cmd_check = f'powershell "Get-Printer -Name \'{n}\' | Select-Object -ExpandProperty Shared"'
                out = subprocess.check_output(cmd_check, shell=True).decode().strip().lower()
                new_state = "$true" if out != "true" else "$false"
                subprocess.run(f'powershell "Set-Printer -Name \'{n}\' -Shared {new_state}"', shell=True)
                self.scan_printers()
            except: pass

    def action_delete(self):
        sel = self.tree.selection()
        if not sel: return
        vals = self.tree.item(sel[0])['values']
        n = vals[2].lstrip("⭐").strip()
        d = vals[4]
        if messagebox.askyesno("XÓA", f"Xóa máy in: {n}?\n(Tự động backup trước khi xóa)"):
            self.run_thread(self.process_delete, args=(n, d))

    def process_delete(self, p_name, d_name):
        self.log(f"Đang xóa: {p_name}")
        self.backup_registry(p_name)
        self.stop_spooler()
        
        path = f"{REG_PRINTERS}\\{p_name}"
        try: self.delete_registry_tree(winreg.HKEY_LOCAL_MACHINE, path)
        except: pass
        
        if d_name and d_name != "N/A":
            for base in [REG_DRIVERS_V3, REG_DRIVERS_V4]:
                try: self.delete_registry_tree(winreg.HKEY_LOCAL_MACHINE, f"{base}\\{d_name}")
                except: pass
        
        self.start_spooler()
        self.log("Xóa hoàn tất.")
        self.root.after(1000, self.scan_printers)

    def backup_registry(self, p_name):
        if not os.path.exists(BACKUP_DIR): os.makedirs(BACKUP_DIR)
        safe = "".join([c for c in p_name if c.isalnum()]).strip()
        fname = os.path.join(BACKUP_DIR, f"{safe}.reg")
        subprocess.run(f'reg export "HKLM\\{REG_PRINTERS}\\{p_name}" "{fname}" /y', shell=True)

    def delete_registry_tree(self, root, path):
        open_key = winreg.OpenKey(root, path, 0, winreg.KEY_ALL_ACCESS)
        info = winreg.QueryInfoKey(open_key)
        for _ in range(info[0]):
            sub = winreg.EnumKey(open_key, 0)
            self.delete_registry_tree(root, f"{path}\\{sub}")
        winreg.CloseKey(open_key)
        winreg.DeleteKey(root, path)

    def delete_driver_reg(self, d_name):
        for base in [REG_DRIVERS_V3, REG_DRIVERS_V4]:
            try: self.delete_registry_tree(winreg.HKEY_LOCAL_MACHINE, f"{base}\\{d_name}")
            except: pass

    def restart_spooler(self):
        self.stop_spooler()
        self.start_spooler()
        self.log("Spooler Restarted.")
        # Thêm thông báo
        messagebox.showinfo("Thành công", "Đã khởi động lại Print Spooler Service!")

    def stop_spooler(self): subprocess.run("net stop spooler", shell=True, creationflags=0x08000000)
    def start_spooler(self): subprocess.run("net start spooler", shell=True, creationflags=0x08000000)

    def clear_spool_files(self):
        if messagebox.askyesno("Xác nhận", "Xóa toàn bộ lệnh in đang chờ (Clear Queue)?"):
            self.stop_spooler()
            try:
                for f in os.listdir(SPOOL_DIR):
                    p = os.path.join(SPOOL_DIR, f)
                    if os.path.isfile(p): os.unlink(p)
                    elif os.path.isdir(p): shutil.rmtree(p)
            except: pass
            self.start_spooler()
            self.log("Đã dọn sạch Queue.")
            messagebox.showinfo("Thành công", "Đã xóa sạch các lệnh in đang treo!")

    def export_report(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            messagebox.showerror("Thiếu thư viện", "Vui lòng chạy lệnh: pip install openpyxl")
            return

        # Tạo tên file
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"Printer_Manager_{timestamp}.xlsx"

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel File", "*.xlsx")],
            title="Xuất Báo Cáo Excel"
        )
        
        if not filename: return
        self.log(f"📥 Đang xuất Excel: {filename}...")

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Printer List"

            headers = ["STT", "Trạng Thái", "Tên Máy In", "Cổng Kết Nối", "Driver", "Chia Sẻ"]
            ws.append(headers)

            # Style Header
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

            # Data
            for item in self.tree.get_children():
                row = self.tree.item(item)['values']
                ws.append(row)

            # Auto fit
            for col in ws.columns:
                max_len = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = (max_len + 2)

            wb.save(filename)
            messagebox.showinfo("Thành công", f"Đã xuất file:\n{filename}")
            self.log("✅ Xuất Excel thành công.")
            os.startfile(filename)

        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
            self.log(f"❌ Lỗi: {e}")

if __name__ == "__main__":
    if is_admin():
        root = tk.Tk()
        app = CleanPrinterApp(root)
        root.mainloop()
    else:
        if run_as_admin(): sys.exit()